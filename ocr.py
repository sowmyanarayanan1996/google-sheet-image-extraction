# ocr.py
import io
import os
import uuid
import requests
import numpy as np

from PIL import Image as PILImage
from openpyxl import load_workbook

# ==========================================
# SAFE OCR IMPORT WITH FALLBACK
# ==========================================

OCR_AVAILABLE = False
engine = None

try:
    from rapidocr_onnxruntime import RapidOCR
    engine = RapidOCR()
    OCR_AVAILABLE = True
    print("✅ RapidOCR loaded successfully")
except Exception as e:
    print(f"⚠️ Could not load RapidOCR: {e}")
    print("📝 OCR functionality will be limited")

# ==========================================
# CREATE TEMP
# ==========================================

os.makedirs("temp/images", exist_ok=True)

# ==========================================
# URL CHECK
# ==========================================

def is_image_url(text):
    if not text:
        return False
    text = str(text).lower()
    return text.startswith("http://") or text.startswith("https://")

# ==========================================
# OCR FUNCTION WITH FALLBACK
# ==========================================

def extract_text_from_image(pil_image):
    if not OCR_AVAILABLE:
        return "[OCR Engine Not Available - Image detected]"
    
    try:
        img_np = np.array(pil_image)
        result, _ = engine(img_np)
        
        if result and len(result) > 0:
            texts = [item[1] for item in result if item and len(item) > 1]
            return "\n".join(texts) if texts else ""
        
        return ""
        
    except Exception as e:
        return f"[OCR Error: {str(e)}]"

# ==========================================
# MAIN FUNCTION
# ==========================================

def process_excel_images(excel_path):
    results = []
    
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # ======================================
        # EMBEDDED IMAGES
        # ======================================
        embedded_images = getattr(ws, "_images", [])
        
        for index, image in enumerate(embedded_images):
            try:
                image_bytes = image._data()
                pil_image = PILImage.open(io.BytesIO(image_bytes)).convert("RGB")
                
                image_name = f"embedded_{index}.png"
                image_path = f"temp/images/{image_name}"
                pil_image.save(image_path)
                
                extracted_text = extract_text_from_image(pil_image)
                
                results.append({
                    "image_name": image_name,
                    "image_path": image_path,
                    "text": extracted_text,
                    "source_type": "embedded_image",
                    "image_url": ""
                })
                
            except Exception as e:
                results.append({
                    "image_name": f"embedded_{index}.png",
                    "image_path": "",
                    "text": "",
                    "source_type": "embedded_image",
                    "image_url": "",
                    "error": str(e)
                })
        
        # ======================================
        # IMAGE URLS
        # ======================================
        for row in ws.iter_rows():
            for cell in row:
                try:
                    cell_value = cell.value
                    
                    if is_image_url(cell_value):
                        image_url = str(cell_value).strip()
                        
                        response = requests.get(image_url, timeout=20)
                        if response.status_code != 200:
                            continue
                        
                        image_bytes = response.content
                        pil_image = PILImage.open(io.BytesIO(image_bytes)).convert("RGB")
                        
                        image_name = f"url_{uuid.uuid4().hex[:8]}.png"
                        image_path = f"temp/images/{image_name}"
                        pil_image.save(image_path)
                        
                        extracted_text = extract_text_from_image(pil_image)
                        
                        results.append({
                            "image_name": image_name,
                            "image_path": image_path,
                            "text": extracted_text,
                            "source_type": "image_url",
                            "image_url": image_url
                        })
                        
                except Exception as e:
                    # Skip individual URL errors
                    continue
                    
    except Exception as e:
        results.append({
            "image_name": "",
            "image_path": "",
            "text": "",
            "source_type": "error",
            "image_url": "",
            "error": f"Workbook Error: {str(e)}"
        })
    
    return results

# import os
# import requests
# import streamlit as st

# from datetime import datetime

# from openpyxl import Workbook
# from openpyxl.drawing.image import Image
# from openpyxl.styles import Alignment

# from database import conn, cursor
# from ocr import process_excel_images

# # ==========================================
# # PAGE CONFIG
# # ==========================================

# st.set_page_config(
#     page_title="Google Sheet OCR",
#     layout="wide"
# )

# # ==========================================
# # CREATE TEMP
# # ==========================================

# os.makedirs(
#     "temp",
#     exist_ok=True
# )

# # ==========================================
# # TITLE
# # ==========================================

# st.title(
#     "Google Sheet Image OCR"
# )

# # ==========================================
# # DASHBOARD COUNTS
# # ==========================================

# cursor.execute("""
# SELECT COUNT(*)
# FROM excel_history
# """)

# total_excels = cursor.fetchone()[0]

# cursor.execute("""
# SELECT COUNT(*)
# FROM ocr_records
# """)

# total_records = cursor.fetchone()[0]

# # ==========================================
# # CARDS
# # ==========================================

# col1, col2 = st.columns(2)

# with col1:

#     st.metric(
#         "Total Excel Processed",
#         total_excels
#     )

# with col2:

#     st.metric(
#         "Total OCR Records",
#         total_records
#     )

# st.markdown("---")

# # ==========================================
# # INPUT
# # ==========================================

# sheet_url = st.text_input(
#     "Enter Google Sheet URL"
# )

# # ==========================================
# # PROCESS
# # ==========================================

# if st.button("Process"):

#     try:

#         # ======================================
#         # VALIDATE
#         # ======================================

#         if "/d/" not in sheet_url:

#             st.error(
#                 "Invalid Google Sheet URL"
#             )

#             st.stop()

#         # ======================================
#         # EXTRACT SHEET ID
#         # ======================================

#         sheet_id = sheet_url.split(
#             "/d/"
#         )[1].split("/")[0]

#         export_url = (
#             f"https://docs.google.com/"
#             f"spreadsheets/d/"
#             f"{sheet_id}/export?format=xlsx"
#         )

#         st.info(
#             "Downloading Excel..."
#         )

#         # ======================================
#         # DOWNLOAD EXCEL
#         # ======================================

#         response = requests.get(
#             export_url
#         )

#         content_type = response.headers.get(
#             "Content-Type",
#             ""
#         )

#         expected = (
#             "application/vnd.openxmlformats-"
#             "officedocument.spreadsheetml.sheet"
#         )

#         if expected not in content_type:

#             st.error(
#                 "Google Sheet must be public"
#             )

#             st.stop()

#         # ======================================
#         # SAVE EXCEL
#         # ======================================

#         excel_path = (
#             "temp/input.xlsx"
#         )

#         with open(
#             excel_path,
#             "wb"
#         ) as f:

#             f.write(
#                 response.content
#             )

#         st.success(
#             "Excel Downloaded"
#         )

#         # ======================================
#         # OCR PROCESS
#         # ======================================

#         with st.spinner(
#             "Running OCR..."
#         ):

#             results = (
#                 process_excel_images(
#                     excel_path
#                 )
#             )

#         st.success(
#             "OCR Completed"
#         )

#         # ======================================
#         # TOTAL RECORDS
#         # ======================================

#         total_excel_records = len(
#             results
#         )

#         # ======================================
#         # STORE EXCEL HISTORY
#         # ======================================

#         created_at = datetime.now().strftime(
#             "%Y-%m-%d %H:%M:%S"
#         )

#         cursor.execute(
#             """
#             INSERT INTO excel_history (

#                 google_sheet_url,
#                 excel_name,
#                 total_records,
#                 created_at

#             )

#             VALUES (?, ?, ?, ?)
#             """,
#             (
#                 sheet_url,
#                 "ocr_output.xlsx",
#                 total_excel_records,
#                 created_at
#             )
#         )

#         conn.commit()

#         excel_history_id = (
#             cursor.lastrowid
#         )

#         # ======================================
#         # OUTPUT EXCEL
#         # ======================================

#         output_wb = Workbook()

#         ws = output_wb.active

#         ws.title = (
#             "OCR Results"
#         )

#         ws["A1"] = "Image"
#         ws["B1"] = "Extracted Text"

#         ws.column_dimensions[
#             "A"
#         ].width = 40

#         ws.column_dimensions[
#             "B"
#         ].width = 80

#         current_row = 2

#         # ======================================
#         # LOOP RESULTS
#         # ======================================

#         for item in results:

#             image_bytes = None

#             # ==================================
#             # IMAGE
#             # ==================================

#             if os.path.exists(
#                 item["image_path"]
#             ):

#                 with open(
#                     item["image_path"],
#                     "rb"
#                 ) as img_file:

#                     image_bytes = (
#                         img_file.read()
#                     )

#                 excel_image = Image(
#                     item["image_path"]
#                 )

#                 excel_image.width = 250
#                 excel_image.height = 180

#                 ws.add_image(
#                     excel_image,
#                     f"A{current_row}"
#                 )

#             # ==================================
#             # TEXT
#             # ==================================

#             ws[f"B{current_row}"] = (
#                 item["text"]
#             )

#             ws[f"B{current_row}"].alignment = Alignment(
#                 wrap_text=True,
#                 vertical="top"
#             )

#             ws.row_dimensions[
#                 current_row
#             ].height = 140

#             current_row += 1

#             # ==================================
#             # STORE DB
#             # ==================================

#             cursor.execute(
#                 """
#                 INSERT INTO ocr_records (

#                     excel_history_id,
#                     image_name,
#                     image_data,
#                     extracted_text,
#                     source_type,
#                     image_url,
#                     created_at

#                 )

#                 VALUES (?, ?, ?, ?, ?, ?, ?)
#                 """,
#                 (
#                     excel_history_id,
#                     item["image_name"],
#                     image_bytes,
#                     item["text"],
#                     item.get(
#                         "source_type",
#                         ""
#                     ),
#                     item.get(
#                         "image_url",
#                         ""
#                     ),
#                     created_at
#                 )
#             )

#         conn.commit()

#         # ======================================
#         # SAVE OUTPUT EXCEL
#         # ======================================

#         output_excel_path = (
#             "temp/ocr_output.xlsx"
#         )

#         output_wb.save(
#             output_excel_path
#         )

#         st.success(
#             f"{total_excel_records} Records Stored"
#         )

#         # ======================================
#         # DOWNLOAD BUTTON
#         # ======================================

#         with open(
#             output_excel_path,
#             "rb"
#         ) as f:

#             st.download_button(

#                 label="Download OCR Excel",

#                 data=f,

#                 file_name="ocr_output.xlsx",

#                 mime=(
#                     "application/vnd.openxmlformats-"
#                     "officedocument.spreadsheetml.sheet"
#                 )
#             )

#     except Exception as e:

#         st.error(
#             f"Error: {str(e)}"
#         )

# # ==========================================
# # HISTORY
# # ==========================================

# st.markdown("---")

# st.subheader(
#     "Processed History"
# )

# cursor.execute("""
# SELECT

#     id,
#     excel_name,
#     total_records,
#     created_at

# FROM excel_history

# ORDER BY id DESC
# """)

# history_rows = cursor.fetchall()

# if history_rows:

#     for row in history_rows:

#         history_id = row[0]

#         with st.expander(
#             f"{row[1]} | Records: {row[2]}"
#         ):

#             st.write(
#                 f"Created At: {row[3]}"
#             )

#             cursor.execute(
#                 """
#                 SELECT

#                     image_name,
#                     image_data,
#                     extracted_text,
#                     source_type

#                 FROM ocr_records

#                 WHERE excel_history_id = ?
#                 """,
#                 (history_id,)
#             )

#             records = cursor.fetchall()

#             for index, rec in enumerate(
#                 records
#             ):

#                 col1, col2 = st.columns(
#                     [1, 2]
#                 )

#                 with col1:

#                     if rec[1]:

#                         st.image(
#                             rec[1],
#                             use_container_width=True
#                         )

#                 with col2:

#                     st.text_area(
#                         f"Extracted Text {index}",
#                         rec[2],
#                         height=150,
#                         disabled=True,
#                         key=f"{history_id}_{index}"
#                     )

# else:

#     st.info(
#         "No history found"
#     )
